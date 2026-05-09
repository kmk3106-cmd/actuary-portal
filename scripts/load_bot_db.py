"""
scripts/load_bot_db.py

Excel 인풋 파일 → actuarial.db 적재 (텔레그램 봇 데이터 소스)

적재 테이블:
  fact_balance      : BEL/RA/CSM/LOSS/OCI 기말잔액 + 잔여보장부채·발생사고부채 합계
  fact_pl           : 보험손익 당월(NP/IDP/VFA/TOTAL) + 누적(TOTAL)
  fact_csm_movement : CSM 무브먼트 코드별 (TOTAL/VFA, 당월)
  metadata          : loaded_periods 업데이트

사용법:
  python scripts/load_bot_db.py --input reports/input/xxx.xlsx --ym 202603
  python scripts/load_bot_db.py --input reports/input/xxx.xlsx --ym 202603 --db server/data/actuarial.db
"""
import argparse
import sqlite3
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string as col_idx
except ImportError:
    print("[오류] openpyxl 미설치: pip install openpyxl")
    sys.exit(1)

# ─────────────────────────────────────────────────────────────
# 상수
# ─────────────────────────────────────────────────────────────

SCHEMA_SQL = '''
CREATE TABLE IF NOT EXISTS metadata (
    key TEXT PRIMARY KEY,
    value TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS fact_balance (
    period_yyyymm TEXT NOT NULL,
    model TEXT NOT NULL,
    metric TEXT NOT NULL,
    amount REAL NOT NULL,
    PRIMARY KEY (period_yyyymm, model, metric)
);
CREATE TABLE IF NOT EXISTS fact_pl (
    period_yyyymm TEXT NOT NULL,
    scope TEXT NOT NULL,
    model TEXT NOT NULL,
    line_item TEXT NOT NULL,
    amount REAL NOT NULL,
    PRIMARY KEY (period_yyyymm, scope, model, line_item)
);
CREATE TABLE IF NOT EXISTS fact_csm_movement (
    start_yyyymm TEXT NOT NULL,
    end_yyyymm TEXT NOT NULL,
    scope TEXT NOT NULL,
    model TEXT NOT NULL,
    movement_code TEXT NOT NULL,
    amount REAL NOT NULL,
    display_order INTEGER NOT NULL DEFAULT 0,
    PRIMARY KEY (start_yyyymm, end_yyyymm, scope, model, movement_code)
);
CREATE TABLE IF NOT EXISTS bot_query_logs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    logged_at_utc TEXT NOT NULL,
    telegram_user_id TEXT,
    username TEXT,
    chat_id TEXT,
    raw_query TEXT NOT NULL,
    parsed_intent TEXT,
    parsed_period TEXT,
    parsed_scope TEXT,
    parsed_metric TEXT,
    parsed_model TEXT,
    status TEXT NOT NULL,
    error_message TEXT
);
'''

# 손익_억 / 회계모형_억 행번호 → (무브먼트 코드, 표시순서)
CSM_MOVEMENT_ROWS = [
    (43, '기말',               0),
    (44, '기시',               1),
    (45, '증감',               2),
    (46, '최초인식',           3),
    (47, '이자비용',           4),
    (48, '경험조정_CSM',       5),
    (49, '경험조정_PL',        6),
    (50, 'RA변동',             7),
    (52, '제거_CSM',           8),
    (53, '모델변경',           9),
    (54, '가정변경_사업비율',  10),
    (55, '가정변경_위험률',    11),
    (56, '가정변경_해지율',    12),
    (57, '가정변경_기타',      13),
    (58, '재량권변경',         14),
    (59, '공시이율예실차',     15),
    (60, '공시이율변경',       16),
    (61, '기타금융가정변경',   17),
    (62, 'VFA기업몫조정',      18),
    (63, 'VFA위험경감',        19),
    (64, '할인율변경',         20),
    (65, 'CSM상각',            21),
    (66, '보험취득CF배분',     22),
    (67, '손실부담비용배분',   23),
    (68, '손실부담비용전환입', 24),
    (69, '이익계약전환추가상각', 25),
]

# 기말잔액 열: E=5(BEL), F=6(RA), G=7(CSM), H=8(LOSS), I=9(OCI)
BALANCE_COLS = {5: 'BEL', 6: 'RA', 7: 'CSM', 8: 'LOSS', 9: 'OCI'}

# 회계모형별 당월 열: D=4(NP), E=5(IDP), F=6(VFA), G=7(TOTAL)
MODEL_COLS = [('NP', 4), ('IDP', 5), ('VFA', 6), ('TOTAL', 7)]

# ─────────────────────────────────────────────────────────────
# 유틸
# ─────────────────────────────────────────────────────────────

def ym_prev(ym: str) -> str:
    y, m = int(ym[:4]), int(ym[4:])
    m -= 1
    if m < 1:
        m, y = 12, y - 1
    return f'{y}{m:02d}'


def safe(val):
    """숫자로 변환 가능한 값만 float 반환, 아니면 None"""
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def find_row_by_label(ws, keywords: list, col: int = 2, start: int = 1, end: int = 80) -> int | None:
    """ws의 지정 열에서 keywords 중 하나가 포함된 첫 번째 행 번호 반환"""
    for r in range(start, end + 1):
        val = str(ws.cell(r, col).value or '').strip()
        for kw in keywords:
            if kw in val:
                return r
    return None


def find_end_row_in_range(ws, label_col: int = 4, start: int = 30, end: int = 120) -> int | None:
    """label_col 열에서 '기말'이 포함된 첫 번째 행 반환"""
    for r in range(start, end + 1):
        if '기말' in str(ws.cell(r, label_col).value or ''):
            return r
    return None


# ─────────────────────────────────────────────────────────────
# DB 초기화 및 조작
# ─────────────────────────────────────────────────────────────

def init_db(db_path: str) -> sqlite3.Connection:
    Path(db_path).parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.executescript(SCHEMA_SQL)
    conn.commit()
    return conn


def upsert_rows(conn: sqlite3.Connection, table: str, cols: list, rows: list) -> None:
    if not rows:
        return
    ph = ','.join(['?'] * len(cols))
    sql = f'INSERT OR REPLACE INTO {table} ({",".join(cols)}) VALUES ({ph})'
    conn.executemany(sql, rows)


def update_loaded_periods(conn: sqlite3.Connection, ym: str) -> None:
    row = conn.execute("SELECT value FROM metadata WHERE key='loaded_periods'").fetchone()
    periods = [x for x in (row[0] if row else '').split(',') if x]
    if ym not in periods:
        periods.append(ym)
    periods_str = ','.join(sorted(set(periods)))
    conn.execute(
        "INSERT INTO metadata(key,value) VALUES('loaded_periods',?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (periods_str,),
    )


# ─────────────────────────────────────────────────────────────
# 데이터 읽기: fact_balance
# ─────────────────────────────────────────────────────────────

def read_balance(wb, ym: str) -> list:
    rows = []

    def extract_balance(ws, model: str, search_start: int, search_end: int) -> None:
        end_r = find_end_row_in_range(ws, label_col=4, start=search_start, end=search_end)
        if not end_r:
            print(f"  [경고] {model} 기말잔액 행 미발견 (검색범위 {search_start}~{search_end})")
            return
        total = 0.0
        for col, metric in BALANCE_COLS.items():
            val = safe(ws.cell(end_r, col).value)
            if val is not None:
                rows.append((ym, model, metric, val))
                total += val
        if total != 0:
            rows.append((ym, model, '잔여보장부채', total))

    # TOTAL: 손익_억 잔여보장부채 기말 (row ~43)
    if '손익_억' in wb.sheetnames:
        ws = wb['손익_억']
        extract_balance(ws, 'TOTAL', 30, 70)

        # 발생사고부채: 두 번째 기말 구간 탐색 (row 70+)
        end_r2 = find_end_row_in_range(ws, label_col=4, start=70, end=120)
        if end_r2:
            bel2 = safe(ws.cell(end_r2, 5).value) or 0
            ra2  = safe(ws.cell(end_r2, 6).value) or 0
            lic  = bel2 + ra2
            if lic != 0:
                rows.append((ym, 'TOTAL', '발생사고부채', lic))
        else:
            print("  [경고] 발생사고부채 기말잔액 행 미발견 (row 70~120)")

    # VFA: 회계모형_억 기말 (row ~43)
    if '회계모형_억' in wb.sheetnames:
        extract_balance(wb['회계모형_억'], 'VFA', 30, 70)

    return rows


# ─────────────────────────────────────────────────────────────
# 데이터 읽기: fact_pl
# ─────────────────────────────────────────────────────────────

def read_pl(wb, ym: str) -> list:
    rows = []

    # ── 당월: 회계모형별 ──
    if '회계모형별' in wb.sheetnames:
        ws = wb['회계모형별']

        # 보험손익_차감전: row 7 = Ⅰ보험손익 (차감전 합계)
        pl_row = find_row_by_label(ws, ['Ⅰ 보험손익', 'Ⅰ보험손익', '보험손익(간접'], col=2, start=5, end=50)
        if pl_row is None:
            pl_row = 7  # 기본값: 주석 확인된 위치
        for model, col in MODEL_COLS:
            val = safe(ws.cell(pl_row, col).value)
            if val is not None:
                rows.append((ym, '당월', model, '보험손익_차감전', val))

        # 간접사업비
        ind_row = find_row_by_label(ws, ['간접사업비'], col=2, start=5, end=60)
        if ind_row:
            for model, col in MODEL_COLS:
                val = safe(ws.cell(ind_row, col).value)
                if val is not None:
                    rows.append((ym, '당월', model, '간접사업비', val))
        else:
            print("  [경고] 회계모형별에서 간접사업비 행 미발견")

        # 보험손익_차감후: 간접사업비 이후 보험손익 행
        if ind_row:
            after_row = find_row_by_label(ws, ['보험손익'], col=2, start=ind_row + 1, end=min(ind_row + 15, 80))
            if after_row:
                for model, col in MODEL_COLS:
                    val = safe(ws.cell(after_row, col).value)
                    if val is not None:
                        rows.append((ym, '당월', model, '보험손익_차감후', val))

    # ── 누적 TOTAL: 손익_요약 ──
    if '손익_요약' in wb.sheetnames:
        ws_sum = wb['손익_요약']
        BE = col_idx('BE')   # 누적 열
        BC = col_idx('BC')   # 당월 열 (교차검증용)

        # 검색할 label → line_item 매핑
        LABEL_MAP = [
            (['Ⅰ   보험손익', 'Ⅰ 보험손익', '보험손익_차감전', '보험손익(간접'], '보험손익_차감전'),
            (['간접사업비'], '간접사업비'),
            (['보험손익_차감후', '간접 차감 후', '간접사업비 차감후'], '보험손익_차감후'),
        ]

        for r in range(5, 55):
            cell_b = str(ws_sum.cell(r, 2).value or '').strip()
            cell_c = str(ws_sum.cell(r, 3).value or '').strip()
            combined = cell_b + ' ' + cell_c

            for keywords, line_item in LABEL_MAP:
                for kw in keywords:
                    if kw in combined:
                        val_be = safe(ws_sum.cell(r, BE).value)
                        if val_be is not None:
                            rows.append((ym, '누적', 'TOTAL', line_item, val_be))
                        break

    return rows


# ─────────────────────────────────────────────────────────────
# 데이터 읽기: fact_csm_movement
# ─────────────────────────────────────────────────────────────

def read_csm_movement(wb, ym: str) -> list:
    rows = []
    prev = ym_prev(ym)
    CSM_COL = 7  # col G = CSM 열

    def read_from_sheet(ws, model: str) -> None:
        for in_row, code, order in CSM_MOVEMENT_ROWS:
            val = safe(ws.cell(in_row, CSM_COL).value)
            if val is not None:
                rows.append((prev, ym, '당월', model, code, val, order))

    if '손익_억' in wb.sheetnames:
        read_from_sheet(wb['손익_억'], 'TOTAL')

    if '회계모형_억' in wb.sheetnames:
        read_from_sheet(wb['회계모형_억'], 'VFA')

    return rows


# ─────────────────────────────────────────────────────────────
# 메인 적재 함수
# ─────────────────────────────────────────────────────────────

def load(input_excel: str, ym: str, db_path: str) -> dict:
    print(f"인풋 파일 로딩: {Path(input_excel).name}")
    wb = openpyxl.load_workbook(input_excel, data_only=True)

    print("데이터 읽기 중...")
    bal_rows = read_balance(wb, ym)
    pl_rows  = read_pl(wb, ym)
    csm_rows = read_csm_movement(wb, ym)

    print(f"DB 적재 중: {db_path}")
    conn = init_db(db_path)

    # 기간 데이터 초기화 (재적재 지원)
    conn.execute("DELETE FROM fact_balance       WHERE period_yyyymm=?", (ym,))
    conn.execute("DELETE FROM fact_pl            WHERE period_yyyymm=?", (ym,))
    conn.execute("DELETE FROM fact_csm_movement  WHERE end_yyyymm=?",    (ym,))

    upsert_rows(conn, 'fact_balance',
                ['period_yyyymm', 'model', 'metric', 'amount'],
                bal_rows)

    upsert_rows(conn, 'fact_pl',
                ['period_yyyymm', 'scope', 'model', 'line_item', 'amount'],
                pl_rows)

    upsert_rows(conn, 'fact_csm_movement',
                ['start_yyyymm', 'end_yyyymm', 'scope', 'model', 'movement_code', 'amount', 'display_order'],
                csm_rows)

    update_loaded_periods(conn, ym)
    conn.commit()
    conn.close()

    return {
        'period': ym,
        'fact_balance': len(bal_rows),
        'fact_pl': len(pl_rows),
        'fact_csm_movement': len(csm_rows),
        'db_path': db_path,
    }


# ─────────────────────────────────────────────────────────────
# CLI 진입점
# ─────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description='Excel 인풋 파일을 actuarial.db에 적재합니다 (텔레그램 봇 데이터 소스).'
    )
    parser.add_argument('--input', required=True,  help='인풋 엑셀 파일 경로 (.xlsx)')
    parser.add_argument('--ym',    required=True,  help='기준 연월 YYYYMM')
    parser.add_argument('--db',    default=None,   help='actuarial.db 경로 (기본: server/data/actuarial.db)')
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent
    db_path = args.db or str(script_dir.parent / 'server' / 'data' / 'actuarial.db')

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"[오류] 인풋 파일 없음: {args.input}")
        sys.exit(1)

    stats = load(str(input_path), args.ym, db_path)

    print()
    print(f"{'=' * 45}")
    print(f"적재 완료: {stats['period']}")
    print(f"  fact_balance      : {stats['fact_balance']:3d} 건")
    print(f"  fact_pl           : {stats['fact_pl']:3d} 건")
    print(f"  fact_csm_movement : {stats['fact_csm_movement']:3d} 건")
    print(f"  DB 경로           : {stats['db_path']}")
    print(f"{'=' * 45}")


if __name__ == '__main__':
    main()
