"""
엑셀 파일의 특정 시트 구조를 점검한다.
지정한 행 범위 안에서 비어있지 않은 셀을 출력한다.

사용법:
    python scripts/check_excel_structure.py --input <엑셀 파일> [--sheet 시트명] [--rows 행수]
"""

import argparse
import sys
import openpyxl


def check_sheet(input_path: str, sheet_name: str, max_rows: int) -> None:
    wb = openpyxl.load_workbook(input_path, data_only=True)

    if sheet_name not in wb.sheetnames:
        print(f"[오류] 시트 '{sheet_name}' 를 찾을 수 없습니다.")
        print(f"  사용 가능한 시트: {', '.join(wb.sheetnames)}")
        sys.exit(1)

    ws = wb[sheet_name]
    max_col = ws.max_column

    print(f"파일  : {input_path}")
    print(f"시트  : {sheet_name}")
    print(f"검사행: 1 ~ {max_rows}  (전체 행: {ws.max_row}, 전체 열: {max_col})")
    print("-" * 70)

    for row_idx in range(1, max_rows + 1):
        row_cells = []
        for col_idx in range(1, max_col + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None and str(val).strip() != "":
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                row_cells.append(f"{col_letter}{row_idx}={repr(val)}")
        if row_cells:
            print(f"  행 {row_idx:3d}: {',  '.join(row_cells)}")

    print("-" * 70)
    print("점검 완료.")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="엑셀 시트 구조 점검 — 비어있지 않은 셀을 행 단위로 출력한다."
    )
    parser.add_argument("--input", required=True, help="점검할 엑셀 파일 경로 (.xlsx)")
    parser.add_argument("--sheet", default="손익", help="점검할 시트명 (기본: 손익)")
    parser.add_argument(
        "--rows", type=int, default=40, help="점검할 행 수 (기본: 40)"
    )
    args = parser.parse_args()

    check_sheet(args.input, args.sheet, args.rows)


if __name__ == "__main__":
    main()
