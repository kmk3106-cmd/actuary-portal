"""
Actuarial Report 엑셀 자동생성 스크립트
사용법: python make_actuarial.py --ym 202604 [--input <xlsx>] [--out <dir>]

채워지는 항목:
  - Sheet 3.보험손익    : 당월 NP/IDP/VFA/Total (회계모형별 시트), 12개월 추세
  - Sheet 8.부채잔액   : 기말/기초/증감 상세 변동 (Total/VFA: 손익_억/회계모형_억)
  - Sheet 9.경험조정   : 당월/누계 경험조정 CSM/PL 항목 (손익_억)
  - Sheet 3 재보험     : 재보험 손익 누계 (손익_요약)
  - 날짜 헤더 전체시트  : 기준월(YYYY.MM) 업데이트
"""
import argparse, glob, json, os, re, shutil, sys
from copy import deepcopy
from datetime import datetime
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string as col_idx
    from openpyxl.styles import PatternFill
except ImportError:
    print(json.dumps({"status":"error","message":"openpyxl 미설치. pip install openpyxl","filename":""}))
    sys.exit(1)

# 테스트용: 인풋에서 가져온 셀 초록 음영 (검증 후 False로 변경)
HIGHLIGHT_INPUT_CELLS = False
GREEN_FILL  = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

# Sheet3 출력행 라벨
SHEET3_ROW_LABELS = {
     7: '보험수익',           8: 'CSM 상각',         9: 'RA 변동',
    10: '보험금',            11: '신계약비',          12: '유지비',
    13: '투자관리비',         14: '손해조사비',        15: '보험취득CF배분',
    16: '손실부담비용배분',   17: '기타',             18: '보험서비스비용',
    19: '발생보험금',         20: '유지비(실제)',      21: '투자관리비(실제)',
    22: '손해조사비(실제)',   23: '손실부담비용전환입', 24: '보험취득CF배분(2)',
    25: '손실부담비용배분(2)',
}

# Sheet8 Total 출력행 라벨
SHEET8_TOTAL_ROW_LABELS = {
     7: '기말',        8: '기초(기시)',    9: '증감',
    10: '최초인식',   11: '이자비용',    12: '경험조정_CSM',  13: '경험조정_PL',
    14: 'RA변동',     15: '제거_CSM',    16: '모델변경',
    17: '가정변경_사업비율', 18: '가정변경_위험률', 19: '가정변경_해지율',
    20: '가정변경_기타', 21: '재량권변경', 22: '공시이율예실차(미래)',
    23: '공시이율변경', 24: '기타금융가정변경', 25: 'VFA기업몫조정',
    26: 'VFA위험경감', 27: '할인율변경', 28: 'CSM상각',
    29: '보험취득CF배분', 30: '손실부담비용배분', 31: '손실부담비용전환입',
    32: '이익계약전환추가상각',
}


def write_missing_sheet(wb_out, missing_log: list, ym: str):
    """누락 항목을 워크북 맨 앞 시트로 삽입한다."""
    from openpyxl.styles import Font, Alignment, Border, Side

    sheet_title = '⚠누락항목'
    if sheet_title in wb_out.sheetnames:
        del wb_out[sheet_title]

    ws = wb_out.create_sheet(title=sheet_title, index=0)

    # 헤더
    headers = ['No', '시트', '항목명', '사유', '인풋 소스']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin

    if not missing_log:
        ws.cell(2, 1, '누락 항목 없음').font = Font(color='007F00')
        ws.column_dimensions['A'].width = 20
        return

    for ri, (sheet, item, reason, source) in enumerate(missing_log, 2):
        for ci, val in enumerate([ri - 1, sheet, item, reason, source], 1):
            cell = ws.cell(ri, ci, val)
            cell.fill = YELLOW_FILL
            cell.border = thin
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = 'A2'

    total = len(missing_log)
    ws.cell(total + 3, 1, f'기준월: {ym[:4]}.{ym[4:]}  /  총 {total}건 누락')
    print(f"누락항목: {total}건 (⚠누락항목 시트 참조)")

# ─────────────────────────────────────────────────────────────
# 유틸
# ─────────────────────────────────────────────────────────────

def ym_to_label(ym: str) -> str:
    """202604 → '26.04"""
    return f"'{ym[2:4]}.{ym[4:]}"

def ym_to_dotted(ym: str) -> str:
    """202604 → 2026.04"""
    return f"{ym[:4]}.{ym[4:]}"

def ym_next(ym: str) -> str:
    """202603 → 202604"""
    y, m = int(ym[:4]), int(ym[4:])
    m += 1
    if m > 12: m, y = 1, y + 1
    return f"{y}{m:02d}"

def ym_prev(ym: str) -> str:
    """202603 → 202602"""
    y, m = int(ym[:4]), int(ym[4:])
    m -= 1
    if m < 1: m, y = 12, y - 1
    return f"{y}{m:02d}"

def find_input_excel(input_dir: str, ym: str):
    for pat in [f"*{ym}*.xlsx", "*.xlsx"]:
        files = [f for f in glob.glob(os.path.join(input_dir, pat))
                 if not os.path.basename(f).startswith('~$')]
        files = sorted(files, key=os.path.getmtime, reverse=True)
        if files:
            return files[0]
    return None

def find_template(output_dir: str, exclude_ym: str = None):
    files = glob.glob(os.path.join(output_dir, "Actuarial Report_*.xlsx"))
    if exclude_ym:
        files = [f for f in files if exclude_ym not in os.path.basename(f)]
    files = sorted(files, key=os.path.getmtime, reverse=True)
    return files[0] if files else None

# ─────────────────────────────────────────────────────────────
# 인풋 데이터 읽기
# ─────────────────────────────────────────────────────────────

# 회계모형별: row6=헤더(구분/NP/IDP/VFA), row7=Ⅰ보험손익, row8=보험수익, row9=CSM상각
# output row 7 ← input row 8, output row 8 ← input row 9, ...
SHEET3_ROW_MAP = {
     7: 8,   # 보험수익
     8: 9,   # CSM 상각
     9: 10,  # RA 변동
    10: 11,  # 보험금
    11: 12,  # 신계약비
    12: 13,  # 유지비
    13: 14,  # 투자관리비
    14: 15,  # 손해조사비
    15: 16,  # 보험취득CF배분
    16: 17,  # 손실부담비용배분
    17: 18,  # 기타
    18: 19,  # 보험서비스비용
    19: 20,  # 발생보험금
    20: 21,  # 유지비(실제)
    21: 22,  # 투자관리비(실제)
    22: 23,  # 손해조사비(실제)
    23: 24,  # 손실부담비용전환입
    24: 25,  # 보험취득CF배분
    25: 26,  # 손실부담비용배분
    # row 37(보험손익 간접차감전)은 계산값, 자동입력 제외
}

# 12개월 추세 (sheet3 rows 58-63) ← 회계모형별 G열(Total당월)
TREND_ROW_MAP = {
    58: 8,   # 보험수익 (input row 8)
    59: 19,  # 보험서비스비용 (input row 19)
    60: None,  # 재보험수익
    61: None,  # 재보험비용
    62: None,  # 간접사업비
    63: None,  # 보험손익(간접차감후) — 계산값
}

def read_hoekemodel(wb_in):
    """회계모형별 시트: 당월 NP/IDP/VFA/Total"""
    ws = wb_in['회계모형별']
    pl = {}   # {row_in: {'NP','IDP','VFA','Total'}}

    for r in range(5, 50):
        d = ws.cell(r, 4).value
        e = ws.cell(r, 5).value
        f = ws.cell(r, 6).value
        g = ws.cell(r, 7).value
        if d is None and e is None and f is None and g is None:
            continue
        pl[r] = {'NP': d, 'IDP': e, 'VFA': f, 'Total': g}

    return pl

def read_sonique_eok(wb_in):
    """손익_억 시트: 잔여보장부채 기말/기초 (BEL/RA/CSM/LOSS/OCI)"""
    ws = wb_in['손익_억']
    bal = {'end': {}, 'beg': {}}
    COLS = {'BEL': 5, 'RA': 6, 'CSM': 7, 'LOSS': 8, 'OCI': 9}
    end_found = beg_found = False
    for r in range(30, 100):
        lbl = ws.cell(r, 4).value
        if lbl and '기말' in str(lbl) and not end_found:
            vals = {k: ws.cell(r, c).value for k, c in COLS.items()}
            if any(v for v in vals.values()):
                bal['end'] = vals; end_found = True
        if lbl and '기초' in str(lbl) and not beg_found:
            vals = {k: ws.cell(r, c).value for k, c in COLS.items()}
            if any(v for v in vals.values()):
                bal['beg'] = vals; beg_found = True
        if end_found and beg_found:
            break
    return bal

def read_sonique_summary(wb_in):
    """손익_요약: 당월 total, 누계 total (회사계 전체, NP/IDP/VFA 구분없음)"""
    ws = wb_in['손익_요약']
    BC = col_idx('BC')
    BE = col_idx('BE')
    cur = {}   # row label → cur total
    ytd = {}
    for r in range(5, 50):
        lb = ws.cell(r, 2).value or ws.cell(r, 3).value
        if not lb: continue
        lb = str(lb).strip()
        cur[lb] = ws.cell(r, BC).value
        ytd[lb] = ws.cell(r, BE).value
    return cur, ytd

def read_sonique_eok_detail(wb_in):
    """
    손익_억 시트: 부채변동 전체 상세 (rows 43~69)
    returns dict: {in_row: [BEL, RA, CSM, LOSS, OCI, 보험서비스비용, 보험수익, 보험금융비용, 보험금융수익]}
    input cols: E=BEL(5), F=RA(6), G=CSM(7), H=LOSS(8), I=OCI(9), J=보험서비스비용(10), K=보험수익(11), L=보험금융비용(12), M=보험금융수익(13)
    """
    ws = wb_in['손익_억']
    result = {}
    for r in range(43, 70):
        row = [ws.cell(r, c).value for c in range(5, 14)]
        if any(v is not None for v in row):
            result[r] = row
    return result

def read_vfa_detail(wb_in):
    """
    회계모형_억 시트: VFA 부채변동 상세 (rows 43~69)
    returns same format as read_sonique_eok_detail
    """
    ws = wb_in['회계모형_억']
    result = {}
    for r in range(43, 70):
        row = [ws.cell(r, c).value for c in range(5, 14)]
        if any(v is not None for v in row):
            result[r] = row
    return result

def read_np_idp_detail(wb_in):
    """
    회계모형_이외 시트: NP+IDP 합산 부채변동 상세 (rows 43~69 원본 = NP+IDP 합계)
    returns same format as read_sonique_eok_detail
    """
    ws = wb_in['회계모형_이외']
    result = {}
    for r in range(43, 70):
        row = [ws.cell(r, c).value for c in range(5, 14)]
        if any(v is not None for v in row):
            # convert from 원 to 억 (divide by 100_000_000)
            converted = []
            for v in row:
                if isinstance(v, (int, float)):
                    converted.append(v / 1e8)
                else:
                    converted.append(v)
            result[r] = converted
    return result

# ─────────────────────────────────────────────────────────────
# Sheet 3 업데이트
# ─────────────────────────────────────────────────────────────

def update_sheet3(ws3, pl_in, cur_total, ytd_total, template_ym: str, new_ym: str, missing_log: list):
    """
    ws3        : 출력 파일의 3.보험손익 시트
    pl_in      : 회계모형별에서 읽은 {row: {NP,IDP,VFA,Total}}
    cur_total  : 손익_요약 당월 합계 dict
    ytd_total  : 손익_요약 누계 합계 dict
    template_ym: 템플릿 기준월 (202603)
    new_ym     : 생성 기준월 (예: 202604)
    """
    # 1) 날짜 헤더 교체
    new_yr = int(new_ym[:4])
    old_dot = ym_to_dotted(template_ym)   # "2026.03"
    new_dot = ym_to_dotted(new_ym)         # "2026.04"

    # 누계 라벨 "2026년 누적" 부분 업데이트
    for row in ws3.iter_rows():
        for cell in row:
            if cell.value == old_dot:
                cell.value = new_dot
            elif isinstance(cell.value, str) and f"{int(new_ym[:4])-1 if int(new_ym[4:])<=3 else int(new_ym[:4])}년 누적" in cell.value:
                pass  # 연도 누계 라벨 업데이트는 아래서 처리
            elif isinstance(cell.value, str) and re.match(r'^\d{4}년 누적$', cell.value):
                cell.value = f"{new_yr}년 누적"

    # 2) 당월 NP/IDP/VFA/Total 업데이트 (col D=4, E=5, F=6, G=7)
    for out_row, in_row in SHEET3_ROW_MAP.items():
        data = pl_in.get(in_row)
        if not data:
            label = SHEET3_ROW_LABELS.get(out_row, f'행{out_row}')
            missing_log.append(('3.보험손익', label, '당월 NP/IDP/VFA/Total 값 없음', f'회계모형별 행{in_row}'))
            continue
        for col, key in [(4,'NP'),(5,'IDP'),(6,'VFA'),(7,'Total')]:
            c = ws3.cell(out_row, col)
            c.value = data[key]
            if HIGHLIGHT_INPUT_CELLS:
                c.fill = GREEN_FILL

    # 3) 누계 Total 업데이트 (col N=14) — NP/IDP/VFA는 미확인 소스라 유지
    _ytd_keys = {
         7: ('Ⅰ.1   보험수익', '보험수익'),
        18: ('Ⅰ.2   보험서비스비용', '보험서비스비용'),
        37: ('Ⅰ   보험손익', '보험손익'),
    }
    for out_row, keys in _ytd_keys.items():
        found = False
        for k in keys:
            val = ytd_total.get(k)
            if val is not None:
                c = ws3.cell(out_row, 14)
                c.value = val
                if HIGHLIGHT_INPUT_CELLS:
                    c.fill = GREEN_FILL
                found = True
                break
        if not found:
            label = SHEET3_ROW_LABELS.get(out_row, f'행{out_row}')
            missing_log.append(('3.보험손익', f'{label} 누계', '손익_요약 누계 합계 값 없음', f'손익_요약 (키: {keys[0]})'))

    # 4) 12개월 추세 업데이트 (rows 57–63)
    # Row 57: 헤더 — 컬럼 C~N (col 3~14)
    # 기존 헤더를 1칸씩 좌이동하고 마지막 컬럼(14)에 새 월 삽입
    hdr_row = 57
    # 헤더 셀 col 3~14 값 읽기
    old_hdrs = [ws3.cell(hdr_row, c).value for c in range(3, 15)]
    # 왼쪽으로 1칸 shift (col3 = col4 old, ..., col13 = col14 old)
    new_hdrs = old_hdrs[1:] + [ym_to_label(new_ym)]
    for c, h in enumerate(new_hdrs, 3):
        ws3.cell(hdr_row, c).value = h

    # 추세 데이터 rows 58~63
    for out_row, in_row in TREND_ROW_MAP.items():
        old_vals = [ws3.cell(out_row, c).value for c in range(3, 15)]
        shifted = old_vals[1:]  # 11개

        if in_row is not None:
            new_val = pl_in.get(in_row, {}).get('Total')
            if new_val is None:
                label = SHEET3_ROW_LABELS.get(in_row, f'행{in_row}')
                missing_log.append(('3.보험손익', f'추세 {label}', '12개월 추세 최신값 없음', f'회계모형별 행{in_row} Total열'))
        else:
            # 재보험/간접사업비는 손익_요약에서 취득 불가 → None 유지
            new_val = None

        shifted.append(new_val)
        for ci, v in enumerate(shifted, 3):
            cell = ws3.cell(out_row, ci)
            cell.value = v
            # 마지막 열(col 14)이 이번 달 새 값 — 초록 표시
            if HIGHLIGHT_INPUT_CELLS and ci == 14 and in_row is not None and v is not None:
                cell.fill = GREEN_FILL

# ─────────────────────────────────────────────────────────────
# Sheet 8 업데이트
# ─────────────────────────────────────────────────────────────

# 출력 Sheet8 row → 입력 손익_억/회계모형_억 row 매핑
# 출력 Section 1 (Total) rows 7~32 → 입력 rows 43~69 (순서 그대로)
SHEET8_TOTAL_ROW_MAP = {
     7: 43,  # 기말
     8: 44,  # 기초(기시)
     9: 45,  # 기중변동(증감)
    10: 46,  # 최초인식
    11: 47,  # 이자비용
    12: 48,  # 경험조정_CSM
    13: 49,  # 경험조정_PL
    14: 50,  # RA변동
    # 15: 51,  # 공시이율예실차_당기 (0값만, skip)
    15: 52,  # 제거_CSM
    16: 53,  # 모델변경
    17: 54,  # 계리가정변경_사업비율
    18: 55,  # 계리가정변경_위험률
    19: 56,  # 계리가정변경_해지율
    20: 57,  # 계리가정변경_기타
    21: 58,  # 재량권변경
    22: 59,  # 공시이율예실차(미래)
    23: 60,  # 공시이율/펀드수익률변경
    24: 61,  # 기타금융가정변경
    25: 62,  # VFA 기업의 몫 조정
    26: 63,  # VFA 위험경감
    27: 64,  # 할인율 변경
    28: 65,  # CSM상각
    29: 66,  # 보험취득현금흐름 배분
    30: 67,  # 손실부담비용 배분
    31: 68,  # 손실부담비용 전/환입
    32: 69,  # 이익계약전환추가상각
}

# 출력 Section 4 (VFA) rows 100~125 → 입력 회계모형_억 rows 43~69 (동일)
SHEET8_VFA_ROW_MAP = {
    100: 43,  # 기말
    101: 44,  # 기초
    102: 45,  # 증감
    103: 46,  # 최초인식
    104: 47,  # 이자비용
    105: 48,  # 경험조정_CSM
    106: 49,  # 경험조정_PL
    107: 50,  # RA변동
    108: 52,  # 제거_CSM
    109: 53,  # 모델변경
    110: 54,  # 계리가정변경_사업비율
    111: 55,  # 계리가정변경_위험률
    112: 56,  # 계리가정변경_해지율
    113: 57,  # 계리가정변경_기타
    114: 58,  # 재량권변경
    115: 59,  # 공시이율예실차(미래)
    116: 60,  # 공시이율/펀드수익률변경
    117: 61,  # 기타금융가정변경
    118: 62,  # VFA 기업의 몫 조정
    119: 63,  # VFA 위험경감
    120: 64,  # 할인율 변경
    121: 65,  # CSM상각
    122: 66,  # 보험취득현금흐름 배분
    123: 67,  # 손실부담비용 배분
    124: 68,  # 손실부담비용 전/환입
    125: 69,  # 이익계약전환추가상각
}

# 출력 col C~K (3~11) → 입력 인덱스 0~8 (BEL,RA,CSM,LOSS,OCI,보서비용,보수익,보금융비용,보금융수익)
SHEET8_COL_START = 3   # output col C
SHEET8_IN_COUNT  = 9   # BEL,RA,CSM,LOSS,OCI,보험서비스비용,보험서비스수익,보험금융비용,보험금융수익


def update_sheet8(ws8, bal_eok, detail_total=None, detail_vfa=None, missing_log=None):
    """
    부채잔액 (Sheet 8) 업데이트
    1. Total 구역 기말/기시/증감 + 전체 상세 행 (손익_억)
    2. VFA 구역 기말/기시/증감 + 전체 상세 행 (회계모형_억)
    3. 기존 기말/기초/증감 BEL/RA/CSM/LOSS/OCI (이전 방식 유지)
    """
    # ── 기존 방식: 동적 탐색으로 기말/기시/증감 (BEL/RA/CSM/LOSS/OCI) ──
    COL = {'BEL': 3, 'RA': 4, 'CSM': 5, 'LOSS': 6, 'OCI': 7}

    def safe_sub(a, b):
        try: return (a or 0) - (b or 0)
        except: return None

    end_row = beg_row = chg_row = None
    for r in range(1, 20):
        lbl = ws8.cell(r, 2).value
        if lbl == '기말' and end_row is None:
            end_row = r
        elif lbl == '기시' and beg_row is None:
            beg_row = r
        elif lbl == '증감' and chg_row is None:
            chg_row = r
        if end_row and beg_row and chg_row:
            break

    if missing_log is None:
        missing_log = []

    end_d = bal_eok.get('end', {})
    beg_d = bal_eok.get('beg', {})

    if not end_d:
        missing_log.append(('8.부채변동', '잔여보장부채 기말', '손익_억에서 기말 행 미발견', '손익_억 행30~100 D열 "기말"'))
    if not beg_d:
        missing_log.append(('8.부채변동', '잔여보장부채 기초', '손익_억에서 기초 행 미발견', '손익_억 행30~100 D열 "기초"'))

    for key, c in COL.items():
        if end_row and key in end_d:
            cell = ws8.cell(end_row, c)
            cell.value = end_d[key]
            if HIGHLIGHT_INPUT_CELLS: cell.fill = GREEN_FILL
        if beg_row and key in beg_d:
            cell = ws8.cell(beg_row, c)
            cell.value = beg_d[key]
            if HIGHLIGHT_INPUT_CELLS: cell.fill = GREEN_FILL
        if chg_row and key in end_d and key in beg_d:
            cell = ws8.cell(chg_row, c)
            cell.value = safe_sub(end_d[key], beg_d[key])
            if HIGHLIGHT_INPUT_CELLS: cell.fill = GREEN_FILL

    # ── 신규: Total 상세 (손익_억) ──
    if detail_total:
        for out_row, in_row in SHEET8_TOTAL_ROW_MAP.items():
            row_data = detail_total.get(in_row)
            if row_data is None:
                label = SHEET8_TOTAL_ROW_LABELS.get(out_row, f'행{out_row}')
                missing_log.append(('8.부채변동', f'Total {label}', '손익_억 해당 행 데이터 없음', f'손익_억 행{in_row}'))
                continue
            for i, val in enumerate(row_data[:SHEET8_IN_COUNT]):
                cell = ws8.cell(out_row, SHEET8_COL_START + i)
                cell.value = val
                if HIGHLIGHT_INPUT_CELLS and val is not None:
                    cell.fill = GREEN_FILL

    # ── 신규: VFA 상세 (회계모형_억) ──
    if detail_vfa:
        for out_row, in_row in SHEET8_VFA_ROW_MAP.items():
            row_data = detail_vfa.get(in_row)
            if row_data is None:
                missing_log.append(('8.부채변동', f'VFA 행{out_row}', '회계모형_억 해당 행 데이터 없음', f'회계모형_억 행{in_row}'))
                continue
            for i, val in enumerate(row_data[:SHEET8_IN_COUNT]):
                cell = ws8.cell(out_row, SHEET8_COL_START + i)
                cell.value = val
                if HIGHLIGHT_INPUT_CELLS and val is not None:
                    cell.fill = GREEN_FILL

# ─────────────────────────────────────────────────────────────
# Sheet 9 업데이트 (경험조정 상세)
# ─────────────────────────────────────────────────────────────

def read_sheet9_data(wb_in):
    """
    손익_억 시트에서 Sheet 9 경험조정 상세 데이터를 읽음.
    Total 당월: 손익_억 R48(경험조정_CSM), R49(경험조정_PL) 상세
    손익_억 R10-R19: 보험수익/서비스비용 상세 항목

    반환: dict with keys 'cur_csm_bel', 'cur_csm_real', 'cur_pl_exp', 'cur_pl_real',
          'cur_csm_sub_bel', 'cur_csm_sub_real' (CSM 세부항목 예상/실제),
          'cur_pl_items_exp', 'cur_pl_items_real' (PL 세부항목 예상/실제)
    """
    ws = wb_in['손익_억']

    def g(r, c): return ws.cell(r, c).value

    data = {}

    # ─ 경험조정 CSM ─
    # R48: 경험조정_CSM: E=BEL조정(예상), G=CSM조정, H=LOSS조정, J=보서비용, K=보수익
    # 실제 CF는 손익_억 R10.1~R11.4 에서 읽음
    # 출력 Sheet9 R8: 1. 경험조정(CSM) 예상=col C, 실제=col D, 차이=col E, CSM=col F, Loss=col G, 보서비용=col H, 보수익=col I

    # R48 cols: E=BEL(-843.35), F=RA(0), G=CSM(-486.89), H=LOSS(117.51), I=OCI, J=보서비용(117.51), K=보수익(0)
    data['r48'] = [g(48, c) for c in range(5, 14)]  # E~M

    # R49 cols: E=BEL(-603.14), G=CSM(0), H=LOSS(0), J=보서비용(740.80), K=보수익(603.14)
    data['r49'] = [g(49, c) for c in range(5, 14)]  # E~M

    # 손익_억 has cumulative data in different location; for now read 당월 only
    # 경험조정 CSM 세부 (수입보험료_CSM, 약관대출_CSM, 신계약비_CSM, 보험금_CSM) 예상/실제
    # 이는 손익_요약 또는 경험조정 시트에서 읽어야 하므로 현재는 스킵
    return data


def read_sheet9_from_soniq_eok(wb_in):
    """손익_억 시트에서 경험조정 Total 항목 읽기"""
    ws = wb_in['손익_억']
    # Row 48: 경험조정_CSM  col E=BEL예상조정, G=CSM조정, H=LOSS조정, J=보험서비스비용, K=보험서비스수익
    # Row 49: 경험조정_PL   col E=BEL예상, J=보험서비스비용, K=보험서비스수익

    def v(r, c): return ws.cell(r, c).value

    # Output Sheet9 R8 cols: B=구분, C=예상(BEL조정), D=실제(현금), E=차이, F=CSM조정, G=Loss조정, H=보서비용, I=보수익
    # R8 = 1. 경험조정(CSM) Total
    #   C = 손익_억 R48.E (경험조정_CSM BEL예상 = -843.35)
    #   F = 손익_억 R48.G (CSM조정 = -486.89)
    #   G = 손익_억 R48.H (Loss조정 = 117.51)
    #   H = 손익_억 R48.J (보서비용 = 117.51)
    #   I = 손익_억 R48.K (보수익 = 0)

    # R14 = 2. 경험조정(PL) Total
    #   C = 손익_억 R49.E (경험조정_PL BEL = -603.14)
    #   H = 손익_억 R49.J (보서비용 = 740.80)
    #   I = 손익_억 R49.K (보수익 = 603.14)

    result = {
        'csm_total': {
            'bel_adj': v(48, 5),    # C
            'csm_adj': v(48, 7),    # F
            'loss_adj': v(48, 8),   # G
            'insv_cost': v(48, 10), # H
            'insv_rev': v(48, 11),  # I
        },
        'pl_total': {
            'bel': v(49, 5),        # C
            'insv_cost': v(49, 10), # H
            'insv_rev': v(49, 11),  # I
        },
        # PL 세부 항목 (손익_억 R13~R18)
        'pl_items': {
            '신계약비':  v(13, 7),   # 손익_억 R13 col G = -10.64
            '유지비':    v(14, 7),   # R14 col G = 72.88
            '투자관리비': v(15, 7),  # R15 col G = 12.20
            '손해조사비': v(16, 7),  # R16 col G = 4.13 (실제: col G is actual)
            '보험금':    v(18, 7),   # R18 col G
        }
    }
    return result


def update_sheet9(ws9, wb_in, missing_log=None):
    """
    Sheet 9 경험조정 상세 업데이트
    - 1. Total 당월 경험조정(CSM)/경험조정(PL) → 손익_억 rows 48, 49
    - 세부 항목: 손익_억 rows 9~19
    """
    if missing_log is None:
        missing_log = []
    ws_eok = wb_in['손익_억']
    def v(r, c): return ws_eok.cell(r, c).value

    # ─── 1. Total 당월 (sheet9 rows 8, 14) ───
    # Row 8: 1. 경험조정(CSM)
    #   col C=예상(BEL조정), col F=CSM조정, col G=Loss조정, col H=보험서비스비용, col I=보험서비스수익
    csm_bel  = v(48, 5)   # BEL 조정 예상
    csm_csm  = v(48, 7)   # CSM 조정
    csm_loss = v(48, 8)   # Loss 조정
    csm_cost = v(48, 10)  # 보험서비스비용
    csm_rev  = v(48, 11)  # 보험서비스수익

    if csm_bel is not None:
        for col, val in [(3, csm_bel),(6, csm_csm),(7, csm_loss),(8, csm_cost),(9, csm_rev)]:
            cell = ws9.cell(8, col)
            cell.value = val
            if HIGHLIGHT_INPUT_CELLS and val is not None:
                cell.fill = GREEN_FILL
    else:
        missing_log.append(('9.경험조정', '경험조정(CSM) Total', '손익_억 행48 E열 값 없음', '손익_억 행48'))

    # Row 14: 2. 경험조정(PL)
    #   col C=예상(BEL조정) ← 손익_억 R49.E 의 부호 반전 (BEL감소 = PL예상 양수)
    #   col H=보험서비스비용, col I=보험서비스수익
    pl_bel_raw = v(49, 5)
    pl_cost    = v(49, 10)
    pl_rev     = v(49, 11)
    # Sign convention: 손익_억 R49.E is negative (BEL decreases), output shows positive expected value
    pl_bel = -pl_bel_raw if isinstance(pl_bel_raw, (int, float)) else pl_bel_raw

    if pl_bel is not None:
        for col, val in [(3, pl_bel),(8, pl_cost),(9, pl_rev)]:
            cell = ws9.cell(14, col)
            cell.value = val
            if HIGHLIGHT_INPUT_CELLS and val is not None:
                cell.fill = GREEN_FILL
    else:
        missing_log.append(('9.경험조정', '경험조정(PL) Total', '손익_억 행49 E열 값 없음', '손익_억 행49'))

    # ─── PL 세부 항목 (rows 15~20) ───
    # 손익_억 R13~R18 col G = 예상(보험수익 입장에서의 예상값=BEL), col H = 실제 아님
    # 실제로는 손익_요약에서 취득 필요하므로, 손익_억 당월 총계 항목 매핑
    # R13: 신계약비 PL예상-실제, R14: 유지비, R15: 투자관리비, R16: 손해조사비, R18: 보험금
    # Output R15~R20: 신계약비, 유지비, 투자관리비, 손해조사비, 과거보험료, 보험금
    # col C=예상, D=실제, E=차이  (단위: 억원)
    # 손익_억 row 10~18: col G=경비관련 당월값들

    # 보험금 예상: 손익_요약 경험조정 행에서 읽음 (R29 보험금_PL 예상)
    # For now, use available data from 손익_억
    # R13.G = 신계약비(예상-실제) = -10.64... R13.J = 예상보험수익, 실제보험수익
    # The 손익_억 sheet has 'SP.4: -10.64419218' as col G, row 13
    # These are already aggregated actual-expected differences, not separate exp/act
    # Better source: 경험조정 sheet for detail

    # From 경험조정 sheet: row 5=총계, rows 10-16 = PL 항목별
    ws_kyung = wb_in['경험조정']
    # R10 CF_PL 예상 = col 3 (C), 실제 = 없음 직접 (각 항목 별도)
    # R11 유지비 예상 = col 4 (D), 실제 = col 5 (E)? Let's check structure

    # 경험조정 sheet R4: 구분 / 예상 / ... / 실제
    # R5: 투자요소 CF_CSM 예상 col B=CF_CSM, col C=BEL예상, col D=0, col E=CSM, ...
    # R10: 보험요소 CF_PL 예상 col C=BEL, J=보험서비스비용, K=보험서비스수익
    # Rows 11-16: 각 항목별 예상 col C, 실제 필요
    # The 'actual' is in 손익_요약 BE column (누계) or 손익_억 당월 항목

    # 실제 = 손익_억 (단위 억)
    # 신계약비 실제: 손익_억 R13 col F (신계약비 보험수익)
    # Actually 손익_억 col G is the expense item value per its structure
    # Let's use 손익_요약 BC col for 당월 actual values (in 억)
    ws_sum = wb_in['손익_요약']
    BC = col_idx('BC')

    # 손익_요약 rows: 신계약비=row 13, 유지비=row 14, 투자관리비=row 15, 손해조사비=row16, 보험금=row 12
    # Need to find the actual row numbers dynamically
    pl_actual = {}
    for r in range(7, 35):
        lbl = ws_sum.cell(r, 3).value
        if not lbl: lbl = ws_sum.cell(r, 2).value
        if not lbl: continue
        lbl = str(lbl).strip()
        val_bc = ws_sum.cell(r, BC).value
        if lbl in ('신계약비', ' 신계약비', 'Ⅰ.2.3   신계약비'):
            pl_actual['신계약비'] = val_bc
        elif lbl in ('유지비', ' 유지비', 'Ⅰ.2.4   유지비'):
            pl_actual['유지비'] = val_bc
        elif '투자관리비' in lbl:
            pl_actual['투자관리비'] = val_bc
        elif '손해조사비' in lbl:
            pl_actual['손해조사비'] = val_bc
        elif lbl in ('보험금', ' 보험금', 'Ⅰ.2.1   발생보험금', '발생보험금'):
            pl_actual['보험금'] = val_bc

    # 손익_억에서 예상(보험서비스비용 항목)  읽기
    # R12: 보험금(보험서비스비용 구성요소) col G = 444.37 (실제, 억)
    # R13: 신계약비 col G = -10.64 (예상-실제)
    # R14: 유지비 col G = 72.88 (예상)
    # 예상값은 손익_억 col G (SP.3,4,5 등)
    # But these are labeled as expense items, not clear exp vs actual separation
    # Use the 손익_요약 BC col for '실제' and compute expected from known formula

    # For Sheet 9 rows 15-20 당월 예상(C), 실제(D), 차이(E):
    # We can get the actual from 손익_요약 BC col
    # Predicted (expected) = actual - difference
    # difference is in 손익_억 R13 (신계약비 차이 = 예상-실제 = -10.64)
    # 신계약비: actual = 손익_요약 BC, expected = actual + 차이(예상-실제)

    # Row 15: 신계약비 (P/L)
    # Output: C=expected, D=actual, E=difference
    eok_diff = {
        '신계약비':  v(13, 7),   # 손익_억 R13 col G = 예상-실제 = -10.64
        '유지비':    v(14, 7),   # R14 = 72.88 (이건 유지비 예상값 자체임)
        '투자관리비': v(15, 7),  # R15
        '손해조사비': v(16, 7),  # R16
    }
    # 손익_억 R12 col G = 보험금 실제값 = 444.37 (not 예상-실제)
    # 손익_억 labels: SP.3=신계약비예상-실제, SP.4=유지비예상, SP.5=투자관리비예상, SP.6=손해조사비예상
    # 실제로 확인: R13 = 신계약비 예상-실제(=BEL조정), R14=유지비, R15=투자관리비, R16=손해조사비
    # 이 값들은 보험서비스비용 입장에서 예상값임
    # 단순히 넣기만 하면 됨 → skip for now, 확실한 source 불분명

    # 경험조정 PL 세부는 데이터가 있으나 손익_억에서 예상/실제 분리가 불명확해 건드리지 않음

    pass  # 세부 항목은 별도 분석 필요, 현재 Total만 채움


# ─────────────────────────────────────────────────────────────
# 전체 날짜 헤더 교체
# ─────────────────────────────────────────────────────────────

def replace_date_headers(wb_out, template_ym: str, new_ym: str):
    old_dot = ym_to_dotted(template_ym)
    new_dot = ym_to_dotted(new_ym)
    old_yr  = template_ym[:4]
    new_yr  = new_ym[:4]

    for sname in wb_out.sheetnames:
        ws = wb_out[sname]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v == old_dot:
                    cell.value = new_dot
                elif isinstance(v, str):
                    if old_dot in v:
                        cell.value = v.replace(old_dot, new_dot)

# ─────────────────────────────────────────────────────────────
# 메인 생성 함수
# ─────────────────────────────────────────────────────────────

def generate(input_excel: str, template_path: str, output_path: str, ym: str):
    print(f"처리중: 인풋 파일 로딩 → {os.path.basename(input_excel)}")
    wb_in = openpyxl.load_workbook(input_excel, data_only=True)

    # 템플릿 기준월 탐색
    m = re.search(r'(\d{6})', os.path.basename(template_path))
    template_ym = m.group(1) if m else "202603"

    print(f"처리중: 템플릿 복사 ({template_ym} → {ym})")
    import tempfile
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.close()
    shutil.copy2(template_path, tmp.name)
    wb_out = openpyxl.load_workbook(tmp.name)
    os.unlink(tmp.name)

    # 데이터 읽기
    pl_in     = read_hoekemodel(wb_in)
    cur_total, ytd_total = read_sonique_summary(wb_in)
    bal_eok   = read_sonique_eok(wb_in)
    detail_total = read_sonique_eok_detail(wb_in)
    detail_vfa   = read_vfa_detail(wb_in)

    missing_log = []

    # Sheet 3 업데이트
    if '3.보험손익' in wb_out.sheetnames:
        print("처리중: Sheet 3 보험손익 업데이트")
        update_sheet3(wb_out['3.보험손익'], pl_in, cur_total, ytd_total, template_ym, ym, missing_log)

    # Sheet 8 업데이트
    if '8.회계모형별 부채 및 관련계정 변동' in wb_out.sheetnames:
        print("처리중: Sheet 8 부채잔액 업데이트")
        update_sheet8(wb_out['8.회계모형별 부채 및 관련계정 변동'], bal_eok, detail_total, detail_vfa, missing_log)

    # Sheet 9 업데이트
    if '9.경험조정 상세' in wb_out.sheetnames:
        print("처리중: Sheet 9 경험조정 상세 업데이트")
        update_sheet9(wb_out['9.경험조정 상세'], wb_in, missing_log)

    # 전체 날짜 헤더 교체
    print("처리중: 날짜 헤더 전체 업데이트")
    replace_date_headers(wb_out, template_ym, ym)

    # 누락항목 시트 (맨 앞)
    write_missing_sheet(wb_out, missing_log, ym)

    wb_out.save(output_path)
    print(f"생성완료: {os.path.basename(output_path)}")
    return output_path


# ─────────────────────────────────────────────────────────────
# CLI 진입점
# ─────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--ym',    required=True, help='기준월 YYYYMM')
    ap.add_argument('--input', default=None,  help='인풋 엑셀 경로 (생략시 자동탐색)')
    ap.add_argument('--out',   default=None,  help='출력 디렉토리 (생략시 reports/output)')
    args = ap.parse_args()

    script_dir = Path(__file__).parent
    input_dir  = script_dir / 'input'
    output_dir = script_dir / 'output'
    if args.out:
        output_dir = Path(args.out)
    output_dir.mkdir(parents=True, exist_ok=True)

    input_excel = args.input or find_input_excel(str(input_dir), args.ym)
    if not input_excel:
        print(json.dumps({"status":"error","message":f"인풋 파일을 찾을 수 없습니다 (ym={args.ym})", "filename":""}))
        sys.exit(1)

    template_path = find_template(str(output_dir))
    if not template_path:
        print(json.dumps({"status":"error","message":"템플릿 파일(Actuarial Report_*.xlsx)이 없습니다","filename":""}))
        sys.exit(1)

    filename = f"Actuarial Report_{args.ym}.xlsx"
    output_path = str(output_dir / filename)

    try:
        generate(input_excel, template_path, output_path, args.ym)
        print(json.dumps({"status":"success","filename":filename,"path":output_path}, ensure_ascii=False))
    except Exception as e:
        import traceback; traceback.print_exc()
        print(json.dumps({"status":"error","message":str(e),"filename":""}, ensure_ascii=False))
        sys.exit(1)

if __name__ == '__main__':
    main()
