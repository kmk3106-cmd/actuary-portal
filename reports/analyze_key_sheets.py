import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import os

filepath = r'C:\Users\USER\actuary potal\reports\input\(이동민)IFRS17재무결산분석_202603_월별.xlsx'

wb_f = openpyxl.load_workbook(filepath, data_only=False)
wb_v = openpyxl.load_workbook(filepath, data_only=True)

output_lines = []

def log(msg=''):
    output_lines.append(msg)
    print(msg)

def dump_sheet(sheet_name, max_row=None, max_col=None):
    ws_f = wb_f[sheet_name]
    ws_v = wb_v[sheet_name]
    log()
    log('=' * 70)
    log(f"시트: [{sheet_name}]")
    log(f"범위: {ws_f.dimensions} (최대행:{ws_f.max_row}, 최대열:{ws_f.max_column})")
    log('=' * 70)

    mr = max_row or ws_f.max_row
    mc = max_col or ws_f.max_column

    for row in ws_f.iter_rows(max_row=mr, max_col=mc):
        for cell in row:
            val = ws_v[cell.coordinate].value
            formula = cell.value
            if val is None and formula is None:
                continue
            if val == '' and (formula is None or formula == ''):
                continue
            has_formula = isinstance(formula, str) and formula.startswith('=')
            if has_formula:
                log(f"  {cell.coordinate}: [수식] {formula}  → 값: {val}")
            elif val is not None and val != '':
                log(f"  {cell.coordinate}: {val}")

# 1. 포트폴리오 (작은 시트)
dump_sheet('포트폴리오')

# 2. 손익_요약 시트 (핵심 보고서용)
dump_sheet('손익_요약')

# 3. 손익 시트 (상세)
dump_sheet('손익')

# 4. 회계모형별 시트
dump_sheet('회계모형별')

# 5. VFA 시트 (앞 50행)
dump_sheet('VFA', max_row=57)

# 6. 발생사고부채
dump_sheet('발생사고부채')

# 7. 발생사고부채_그룹
dump_sheet('발생사고부채_그룹')

# 8. Setting_Report (앞 50행)
dump_sheet('Setting_Report', max_row=50)

# 9. 이자비용
dump_sheet('이자비용')

# 10. 경험조정
dump_sheet('경험조정')

# 11. 예실차
dump_sheet('예실차')

# 12. 예실차_C
dump_sheet('예실차_C')

# 13. CSM변동_VFA
dump_sheet('CSM변동_VFA')

# 14. CSM변동_VFA외
dump_sheet('CSM변동_VFA외')

# 15. 경험조정CSM
dump_sheet('경험조정CSM')

# 16. 기시손실그룹여부 (앞 20행)
dump_sheet('기시손실그룹여부', max_row=20)

# 17. 코드 시트 (앞 50행)
dump_sheet('코드', max_row=50)

# 18. 모델 시트 (앞 10행만)
dump_sheet('모델', max_row=10)

# 저장
out_path = r'C:\Users\USER\actuary potal\reports\key_sheets_analysis.txt'
with open(out_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(output_lines))
log(f"\n저장 완료: {out_path}")
