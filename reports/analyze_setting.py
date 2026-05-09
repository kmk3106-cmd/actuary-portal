import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

filepath = r'C:\Users\USER\actuary potal\reports\input\(이동민)IFRS17재무결산분석_202603_월별.xlsx'

wb_f = openpyxl.load_workbook(filepath, data_only=False)
wb_v = openpyxl.load_workbook(filepath, data_only=True)

output_lines = []
def log(msg=''):
    output_lines.append(str(msg))
    print(msg)

# Setting_Report 시트 전체 상세 분석
sheet_name = 'Setting_Report'
ws_f = wb_f[sheet_name]
ws_v = wb_v[sheet_name]

log('=' * 70)
log(f"시트: [{sheet_name}]")
log(f"범위: {ws_f.dimensions} (최대행:{ws_f.max_row}, 최대열:{ws_f.max_column})")
log('=' * 70)

for row in ws_f.iter_rows(max_row=ws_f.max_row):
    for cell in row:
        val = ws_v[cell.coordinate].value
        formula = cell.value
        if val is None and formula is None:
            continue
        if val == '' and (formula is None or formula == ''):
            continue
        has_formula = isinstance(formula, str) and formula.startswith('=')
        if has_formula:
            log(f"  {cell.coordinate}: [수식] {formula}  → 값: {val}")
        elif val is not None and val != '':
            log(f"  {cell.coordinate}: {val}")

out_path = r'C:\Users\USER\actuary potal\reports\setting_report_analysis.txt'
with open(out_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(output_lines))
log(f"\n저장 완료: {out_path}")
