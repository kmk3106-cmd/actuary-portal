import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import os

filepath = r'C:\Users\USER\actuary potal\reports\input\(이동민)IFRS17재무결산분석_202603_월별.xlsx'

wb_f = openpyxl.load_workbook(filepath, data_only=False)
wb_v = openpyxl.load_workbook(filepath, data_only=True)

print("=" * 80)
print("파일: (이동민)IFRS17재무결산분석_202603_월별.xlsx")
print(f"총 시트 수: {len(wb_f.sheetnames)}")
print("=" * 80)

# 각 시트별 요약 정보
for sheet_name in wb_f.sheetnames:
    ws_f = wb_f[sheet_name]
    ws_v = wb_v[sheet_name]

    # 데이터가 있는 셀 개수
    total_cells = 0
    formula_cells = 0
    value_cells = 0
    sample_headers = []

    for row in ws_f.iter_rows(min_row=1, max_row=min(5, ws_f.max_row)):
        for cell in row:
            if cell.value is not None and cell.value != '':
                if row[0].row == 1:
                    sample_headers.append(str(cell.value)[:20])
                total_cells += 1
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_cells += 1
                else:
                    value_cells += 1

    print(f"\n[{sheet_name}]")
    print(f"  범위: {ws_f.dimensions} (행:{ws_f.max_row}, 열:{ws_f.max_column})")
    print(f"  1행 헤더샘플: {sample_headers[:8]}")
