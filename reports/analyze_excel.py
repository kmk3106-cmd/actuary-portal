import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import os
import glob

# 파일 찾기
input_dir = r'C:\Users\USER\actuary potal\reports\input'
files = glob.glob(os.path.join(input_dir, '*.xlsx')) + glob.glob(os.path.join(input_dir, '*.xls'))

output_lines = []

def log(msg=''):
    output_lines.append(msg)
    print(msg)

log("발견된 파일: " + str(files))

for filepath in files:
    log()
    log('='*80)
    log(f"파일: {os.path.basename(filepath)}")
    log('='*80)

    # 수식 포함해서 읽기
    wb_formula = openpyxl.load_workbook(filepath, data_only=False)
    # 값만 읽기
    wb_value = openpyxl.load_workbook(filepath, data_only=True)

    log(f"시트 목록: {wb_formula.sheetnames}")

    for sheet_name in wb_formula.sheetnames:
        ws_f = wb_formula[sheet_name]
        ws_v = wb_value[sheet_name]

        log()
        log('─'*60)
        log(f"시트: [{sheet_name}]")
        log(f"범위: {ws_f.dimensions}  (최대행:{ws_f.max_row}, 최대열:{ws_f.max_column})")
        log('─'*60)

        for row in ws_f.iter_rows():
            for cell in row:
                val = ws_v[cell.coordinate].value
                formula = cell.value

                # 빈 셀 스킵
                if val is None and formula is None:
                    continue
                if val == '' and (formula is None or formula == ''):
                    continue

                has_formula = isinstance(formula, str) and formula.startswith('=')

                if has_formula:
                    log(f"  {cell.coordinate}: [수식] {formula}  → 값: {val}")
                elif val is not None and val != '':
                    log(f"  {cell.coordinate}: {val}")

# 결과 저장
output_path = r'C:\Users\USER\actuary potal\reports\analysis_result.txt'
with open(output_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(output_lines))

print(f"\n\n분석 결과가 저장되었습니다: {output_path}")
