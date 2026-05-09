"""
엑셀에서 업무명(C열)과 담당자(F열)를 함께 읽어 매핑 확인
"""
import argparse
import openpyxl

parser = argparse.ArgumentParser(description='업무명+담당자 매핑 확인')
parser.add_argument('--input', required=True, help='엑셀 파일 경로')
parser.add_argument('--sheet', default=None, help='시트명 (생략 시 첫 번째 시트)')
parser.add_argument('--out', required=True, help='결과 저장 텍스트 파일')
args = parser.parse_args()

wb = openpyxl.load_workbook(args.input, data_only=True)
ws = wb[args.sheet] if args.sheet else wb.worksheets[0]

lines = [f'시트: {ws.title}', '=' * 60]
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    idx = row[0].row
    # A~H열 전체 출력 (매핑 파악용)
    vals = [str(c.value or '').strip() for c in row[:8]]
    if any(v for v in vals):
        lines.append(f'Row{idx:02d} | ' + ' | '.join(f'{chr(65+i)}={v}' for i, v in enumerate(vals)))

with open(args.out, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f'저장 완료: {args.out}')
